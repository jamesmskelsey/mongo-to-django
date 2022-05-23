"""
Database (.xlsx) to Pydantic model. Assumes the data you want 
to convert is correctly formatted in database.xlsx.

Will take the "Field Name" and "Type" columns on the far left
from database.xlsx 
"""
from dataclasses import field
import json
from openpyxl import load_workbook
from py_to_json import translate_py_to_json

# Open the file


wb = load_workbook(filename="database.xlsx")

print(wb.sheetnames)

def model_file_name(model_name):
    model_name = model_name.lower()
    return f"{model_name}.json"

def capitalize_first(name):
    lower = name.lower()
    return f"{lower[0:1].upper()}{lower[1:]}"

def define_sub(schema, field_name, fields):
    print(field_name, fields)
    for (sub_field_name, sub_field_type) in fields:
        schema["definitions"][field_name]["properties"][sub_field_name[len(f"{field_name}."):]] = {"type": translate_py_to_json(sub_field_type)} 

def singularize(plural):
    # companies -> Company
    plural = capitalize_first(plural).replace("ies", "y")
    # dogs -> Dog
    plural = capitalize_first(plural).replace("s", "")
    return plural


for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    model_as_dict = {}
    for row in ws.values:
        # read the pairs and create a dictionary to make it easy to work with
        field_name = row[0]
        field_type = row[1]
        if field_name != 'Field Name' and field_name != '_id':
            model_as_dict[field_name] = field_type

    schema = {
        "$id": model_file_name(sheetname),
        "$schema": "http://json-schema.org/draft-07/schema#",
        "title": capitalize_first(sheetname),
        "type": "object",
        "properties": {

        },
        "required": [

        ],
        "definitions": {

        }
    }

    # separate out the primary fields (as in the fields that don't belong to objects in an array)
    only_primary_fields = [(e, v) for (e,v) in model_as_dict.items() if "." not in str(e)]
    for field_name, field_type in only_primary_fields:
        # sometimes field name is null, so skip those
        print(field_name)
        if field_name != "null" and field_name != None:
            # but otherwise, we'll just put the type and the name in the properties
            schema['properties'][field_name] = {
                "type": translate_py_to_json(field_type),
                "description": "Replace me."
            }

        if field_type == "array":
            singled_name = singularize(field_name)
            schema['properties'][field_name]["items"] = [{
                "$ref": f"#/definitions/{singled_name}"
            }]
            schema["definitions"] = {
                singled_name: {
                    "properties": {}
                }
            }
            print("Attaching array...")
            define_sub(schema, singled_name, [(e, v) for (e,v) in model_as_dict.items() if f"{field_name}." in str(e)])


    

    # Create a .py file, make the model
    with open(model_file_name(sheetname), 'w', encoding='utf-8') as f:
        f.write(json.dumps(schema, indent=4))
        f.close()
