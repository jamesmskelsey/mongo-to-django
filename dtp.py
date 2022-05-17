"""
Database (.xlsx) to Pydantic model. Assumes the data you want 
to convert is correctly formatted in database.xlsx.

Will take the "Field Name" and "Type" columns on the far left
from database.xlsx 
"""
from dataclasses import field
from openpyxl import load_workbook
# Open the file
wb = load_workbook(filename="database.xlsx")

print(wb.sheetnames)

def model_file_name(model_name):
    model_name = model_name.lower()
    return f"{model_name[0].upper()}{model_name[1:]}.py"

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    model_as_dict = {}
    for row in ws.values:
        # read the pairs and create a dictionary to make it easy to work with
        field_name = row[0]
        field_type = row[1]
        if field_name != 'Field Name' and field_name != '_id':
            model_as_dict[field_name] = field_type
    # spit out the info
    print(model_as_dict)


    # Create a .py file, make the model
    with open(model_file_name(sheetname), 'w', encoding='utf-8') as f:
        for field, f_type in model_as_dict.items():
            f.write(f"{field}: {f_type}\n")
        f.close()

